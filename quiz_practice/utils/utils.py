"""Parse xlsx."""
import datetime
import json
import os

import openpyxl

G2S = {"文学＆歴史": "stage1", "自然科学": "stage2", "現代社会＆地理": "stage3", "グルメ＆趣味": "stage4", "アニメ＆ゲーム": "stage5"}
STAGE_CODE = {"stage1": "1", "stage2": "2", "stage3": "3", "stage4": "4", "stage5": "5"}

# pos_column, pos_row indicate "Q" position
Q_START_POS_COLUMN = 2  # B
Q_START_POS_ROW = 4
QUIZ_POS_ROW_SEEK = 1  # quiz sentence / pos_row + QUIZ_ROW_SEEK
A_POS_COLUMN_SEEK = 1  # "A" / pos_column + A_POS_COLUMN_SEEK
CIRCLE_POS_COLUMN_SEEK = 1  # "⭕" / pos_column + CIRCLE_POS_COLUMN_SEEK
CIRCLE_POS_ROW_SEEK = 1  # "⭕" / pos_column + CIRCLE_POS_ROW_SEEK
CHOICE1_POS_COLUMN_SEEK = 2  # choice1 / pos_column + CHOICE1_POS_COLUMN_SEEK
CHOICE1_POS_ROW_SEEK = 1  # choice1 / pos_row + CHOICE1_POS_ROW_SEEK
Q_NEXT_COLUMN_SEEK = 4  # next "Q" pos /pos_column + Q_NEXT_COLUMN_SEEK
Q_NEXT_ROW_SEEK = 6  # next "Q" pos / pos_row + Q_NEXT_ROW_SEEK
Q_MAX_COLUMN = 10  # J


def parse_xlsx(xlsx_path: str, save_path: str) -> None:
    """Parse xlsx."""
    # init to store data
    quiz_data = {}
    for genre in G2S.keys():
        quiz_data[G2S[genre]] = {"quiz_list": []}

    # parse a xlsx file
    print(f"{xlsx_path} を読込中...")
    workbook = openpyxl.load_workbook(xlsx_path, data_only=True)
    for genre in G2S.keys():
        # check genre existence
        try:
            sheet = workbook[genre]
        except KeyError:
            print(f"{genre} is not Found.")
            continue

        # load quiz
        count = 0
        pos_row = Q_START_POS_ROW
        is_finish = False
        quiz_list = []
        while not is_finish:
            for pos_column in range(Q_START_POS_COLUMN, Q_MAX_COLUMN + 1, Q_NEXT_COLUMN_SEEK):
                skip = False
                Q_cell = sheet.cell(column=pos_column, row=pos_row)
                if Q_cell.value == "Q":
                    A_cell = sheet.cell(column=pos_column + A_POS_COLUMN_SEEK, row=pos_row)
                    circle_cell = sheet.cell(
                        column=pos_column + CIRCLE_POS_COLUMN_SEEK, row=pos_row + CIRCLE_POS_ROW_SEEK
                    )
                    assert A_cell.value == "A"
                    assert circle_cell.value == "⭕"
                    qid = f"{STAGE_CODE[G2S[genre]]}{count+1:05d}"
                    quiz_cell = sheet.cell(column=pos_column, row=pos_row + QUIZ_POS_ROW_SEEK)
                    choice_cell_value_list: list[str] = ["" for _ in range(4)]
                    for i in range(4):
                        choice_cell = sheet.cell(
                            column=pos_column + CHOICE1_POS_COLUMN_SEEK, row=pos_row + CHOICE1_POS_ROW_SEEK + i
                        )
                        choice_cell_value = choice_cell.value
                        if isinstance(choice_cell_value, datetime.datetime):
                            if qid == "300145":
                                choice_cell_value = choice_cell_value.strftime("%#m月%#d日")
                            elif qid == "500176":
                                choice_cell_value = choice_cell_value.strftime("%#m月%#d日")
                            else:
                                skip = True
                        elif not isinstance(choice_cell_value, str):
                            if qid in ["200015", "200037", "200133", "200138", "300060", "400117", "400122", "400150"]:
                                choice_cell_value = str(int(choice_cell_value))
                            else:
                                choice_cell_value = str(int(choice_cell_value))
                                print("Warning!")
                                print(f"Choice of qid={qid} changed ({choice_cell_value} -> {choice_cell.value})")
                        choice_cell_value_list[i] = choice_cell_value

                    quiz = {
                        "qid": qid,
                        "genre": genre,
                        "quiz": quiz_cell.value,
                        "answer": choice_cell_value_list[0],
                        "choices": choice_cell_value_list,
                    }
                    if skip:
                        print("Warning!!!")
                        print(f"Loading qid={qid} is skipped.")
                    else:
                        quiz_list.append(quiz)
                    count += 1
                else:
                    is_finish = True
            pos_row += Q_NEXT_ROW_SEEK
        quiz_data[G2S[genre]] = {"quiz_list": quiz_list}

    # save quiz data
    save_dir = os.path.dirname(save_path)
    if not os.path.exists(save_dir):
        os.makedirs(save_dir, exist_ok=True)
    with open(save_path, mode="w", encoding="utf-8") as f:
        json.dump(quiz_data, f, indent=4, ensure_ascii=False)
